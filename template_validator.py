from openpyxl import load_workbook

REQUIRED_SHEETS = [
    "Prisma Export - Paste as values",
    "Traffic_Doc",
    "Multi-Ad or Creative Rotation"
]

def validate_template(path):
    wb = load_workbook(path, keep_vba=True)

    missing = []

    for sheet in REQUIRED_SHEETS:
        if sheet not in wb.sheetnames:
            missing.append(sheet)

    if missing:
        return False, f"Missing sheets: {', '.join(missing)}"

    return True, "Template Valid"
