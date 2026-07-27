import os
import re

from openpyxl import load_workbook


TRACKING_FILE_NAME = "Pulte_Adobe_Tracking_Codes.xlsm"
TRACKING_SHEET_NAME = "ChannelTrackingValues"


def load_tracking_codes():
    """
    Load Adobe tracking values dynamically from the
    ChannelTrackingValues worksheet.
    """

    file_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        TRACKING_FILE_NAME,
    )

    if not os.path.exists(file_path):
        raise FileNotFoundError(
            f"{TRACKING_FILE_NAME} was not found beside pulte_vip.py."
        )

    workbook = load_workbook(
        file_path,
        data_only=True,
        read_only=True,
    )

    if TRACKING_SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        raise ValueError(
            f"Worksheet '{TRACKING_SHEET_NAME}' was not found "
            f"in {TRACKING_FILE_NAME}."
        )

    worksheet = workbook[TRACKING_SHEET_NAME]

    lookup = {
        "Medium": {},
        "Source": {},
        "Division": {},
        "Region": {},
        "Content": {},
        "Campaign": {},
        "Vendor": {},
        "Image": {},
    }

    columns = {
        "Medium": (1, 2),
        "Source": (3, 4),
        "Division": (5, 6),
        "Region": (7, 8),
        "Content": (9, 10),
        "Campaign": (11, 12),
        "Vendor": (13, 14),
        "Image": (15, 16),
    }

    for category, (name_column, code_column) in columns.items():
        for row_number in range(1, worksheet.max_row + 1):
            name = worksheet.cell(
                row=row_number,
                column=name_column,
            ).value

            code = worksheet.cell(
                row=row_number,
                column=code_column,
            ).value

            if name is None or code is None:
                continue

            clean_name = str(name).strip()
            clean_code = str(code).strip()

            if not clean_name or not clean_code:
                continue

            lookup[category][clean_name.lower()] = clean_code

    workbook.close()

    return lookup


def parse_pulte_vip_placement(placement_name):
    """
    Parse a Pulte VIP Prisma placement name.

    Example:
    Direct_Display_Native_Realtor_Prospect_Geo_DMA_
    2nd Party Data_CPM_Cross Device_1200x800_SS_Local_
    Conversion_EXTD_West Florida_Pulte_Heavy Up_
    Caldera_211086_TAM FL_Onsite
    """

    data = {
        "source": "",
        "dimension": "",
        "image": "",
        "division": "",
        "builder": "",
        "campaign": "",
        "community": "",
        "community_id": "",
        "region": "",
        "placement_type": "",
    }

    if placement_name is None:
        return data

    placement_text = str(placement_name).strip()

    if not placement_text:
        return data

    parts = [
        part.strip()
        for part in placement_text.split("_")
        if part.strip()
    ]

    # Find Source
    for part in parts:
        if part.lower() in {"realtor", "zillow"}:
            data["source"] = part
            break

    # Find Dimension
    for part in parts:
        if re.fullmatch(
            r"\d{2,4}\s*[xX×*]\s*\d{2,4}",
            part,
        ):
            data["dimension"] = (
                part.lower()
                .replace(" ", "")
                .replace("×", "x")
                .replace("*", "x")
            )
            break

    # Find Community ID and surrounding placement values
    for index, part in enumerate(parts):
        if not re.fullmatch(r"\d{5,7}", part):
            continue

        data["community_id"] = part

        if index >= 1:
            data["community"] = parts[index - 1]

        if index >= 2:
            data["campaign"] = parts[index - 2]

        if index >= 3:
            data["builder"] = parts[index - 3]

        if index >= 4:
            data["division"] = parts[index - 4]

        if index >= 5:
            data["image"] = parts[index - 5]

        if index + 1 < len(parts):
            data["region"] = parts[index + 1]

        if index + 2 < len(parts):
            data["placement_type"] = parts[index + 2]

        break

    return data
