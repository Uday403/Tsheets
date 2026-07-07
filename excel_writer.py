import pandas as pd
from openpyxl.cell.cell import MergedCell

def clear_sheet(ws):
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None

def paste_dataframe(ws, df):
    clear_sheet(ws)
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        if not isinstance(cell, MergedCell):
            cell.value = col_name

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if not isinstance(cell, MergedCell):
                cell.value = "" if pd.isna(value) else value

def clear_rows(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None
