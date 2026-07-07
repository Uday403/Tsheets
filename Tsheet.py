import streamlit as st
import pandas as pd
import zipfile
import io
import os
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

st.set_page_config(page_title="Tsheet Automation Dashboard", layout="wide")
st.title("Tsheet Automation Dashboard")

MASTER_TEMPLATE_PATH = "master_template.xlsm"
TRAFFIC_START_ROW = 9

ACCOUNT_TAXONOMY = {
    "Simon": ["Placement Name = Ad Name"],
    "Best Friends Animal Society": ["Placement Name = Ad Name"],
    "BFAS": ["Placement Name = Ad Name"],
    "ConEd": ["Placement Name = Ad Name"],
    "HMH": ["Placement Name = Ad Name"],
    "Ascensus": ["Placement Name = Ad Name"],
    "IMC": ["Placement Name = Ad Name"],
    "Tillamook": ["Placement Name = Ad Name"],
    "AAA": ["Placement Name = Ad Name"],
    "Hyatt": ["Placement Name = Ad Name"],
    "Famous Footwear": ["Placement Name = Ad Name"],
    "Fossil": ["Placement Name = Ad Name"],
    "Lenovo": ["Placement Name = Ad Name"],
    "UPS Store": ["Creative Name = Ad Name"],
    "USTA": ["Dimension = Ad Name"],
    "Pulte": ["Pulte: Market_Brand_Initiative_Property_Duration"],
    "Touchstone Energy": ["Last Meaningful Placement Segment"],
    "Anthem / Elevance": ["ELV_LOB_State_Channel_SizeOrDuration"],
}

STATE_MAP = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
    "FL": "Florida", "GA": "Georgia", "IA": "Iowa", "ID": "Idaho",
    "IL": "Illinois", "IN": "Indiana", "KS": "Kansas", "KY": "Kentucky",
    "LA": "Louisiana", "MA": "Massachusetts", "MD": "Maryland",
    "ME": "Maine", "MI": "Michigan", "MN": "Minnesota", "MO": "Missouri",
    "MS": "Mississippi", "NC": "North Carolina", "NJ": "New Jersey",
    "NV": "Nevada", "NY": "New York", "OH": "Ohio", "PA": "Pennsylvania",
    "TN": "Tennessee", "TX": "Texas", "VA": "Virginia", "WA": "Washington",
    "WI": "Wisconsin", "DC": "District of Columbia"
}

prisma_file = st.file_uploader("Upload Prisma / Tsheet Export", type=["xlsx", "xlsm", "xls", "csv"])

selected_account = st.selectbox("Select Account", list(ACCOUNT_TAXONOMY.keys()))
selected_taxonomy = st.selectbox("Select Ad Name Naming Convention", ACCOUNT_TAXONOMY[selected_account])

creative_type = st.radio("Creative setup type", ["Single creative per ad", "Multiple creatives per ad"])

creative_uploads = st.file_uploader(
    "Upload Creative ZIP or Individual Creative Files",
    type=["zip", "jpg", "jpeg", "png", "gif", "html", "htm", "mp4", "webp"],
    accept_multiple_files=True
)

custom_1x1 = st.text_input("1x1 Creative Name", value="Tracking 1x1")

utm_text = st.text_area(
    "Paste UTM data from Excel. Recommended columns: Placement Name, Dimension, UTM",
    height=150
)

utm_output_column = st.text_input("Traffic_Doc UTM output column", value="P")


def read_prisma_file(uploaded_file):
    ext = uploaded_file.name.split(".")[-1].lower()
    if ext == "csv":
        return pd.read_csv(uploaded_file)
    return pd.read_excel(uploaded_file, sheet_name=0)


def read_utm_text(text):
    if not text.strip():
        return pd.DataFrame()
    try:
        return pd.read_csv(io.StringIO(text), sep="\t")
    except Exception:
        try:
            return pd.read_csv(io.StringIO(text))
        except Exception:
            return pd.DataFrame()


def extract_creative_names(files):
    names = []
    if not files:
        return names

    for uploaded_file in files:
        ext = uploaded_file.name.split(".")[-1].lower()

        if ext == "zip":
            with zipfile.ZipFile(uploaded_file, "r") as zip_ref:
                for file_name in zip_ref.namelist():
                    if not file_name.endswith("/"):
                        names.append(os.path.basename(file_name))
        else:
            names.append(uploaded_file.name)

    return names


def normalize_text(value):
    if pd.isna(value):
        return ""
    return str(value).lower().strip()


def normalize_dimension(value):
    if pd.isna(value):
        return ""
    return str(value).lower().replace(" ", "").replace("*", "x").replace("×", "x")


def split_placement(name):
    if pd.isna(name):
        return []
    name = str(name).strip()
    if "_" in name:
        return [x.strip() for x in name.split("_") if x.strip()]
    if "-" in name:
        return [x.strip() for x in name.split("-") if x.strip()]
    return [name]


def generate_pulte_ad_name(placement_name):
    parts = split_placement(placement_name)

    duration = ""
    for p in parts:
        if p in [":06", ":15", ":30"]:
            duration = p.replace(":", "") + "Sec"

    if "VIID" in parts:
        idx = parts.index("VIID")
        if len(parts) > idx + 4:
            market = parts[idx + 1]
            brand = parts[idx + 2]
            initiative = parts[idx + 3]
            property_name = parts[idx + 4]
            return f"{market}_{brand}_{initiative}_{property_name}_{duration}".strip("_")

    return str(placement_name)


def generate_touchstone_ad_name(placement_name):
    parts = split_placement(placement_name)
    ignore_values = {
        "TEC", "TEC2023", "PGR", "VID", "NAN", "Trade Desk",
        "GM", "Spot X PMP", "dCPM", "CTV", ":15", ":30", ":06", "Third Party"
    }

    meaningful = [p for p in parts if p and p not in ignore_values]
    return meaningful[-1] if meaningful else str(placement_name)


def generate_elv_ad_name(placement_name, dimension):
    parts = split_placement(placement_name)
    text = str(placement_name).lower()

    lob = ""
    state_name = ""
    channel = ""
    size_or_duration = ""

    for p in parts:
        upper = p.upper()
        if upper in ["MDCD", "MDCR", "CSBD", "BRAN"]:
            lob = upper
        if upper in STATE_MAP:
            state_name = STATE_MAP[upper]

    if "display" in text or "banner" in text:
        channel = "Display"
        size_or_duration = str(dimension).strip()
    elif "ctv" in text:
        channel = "CTV"
    elif "olv" in text or "video" in text:
        channel = "OLV"
    elif "audio" in text:
        channel = "Audio"

    if channel != "Display":
        if ":30" in text:
            size_or_duration = "30Sec"
        elif ":15" in text:
            size_or_duration = "15Sec"
        elif ":06" in text:
            size_or_duration = "6Sec"
        else:
            size_or_duration = str(dimension).strip()

    return "_".join(["ELV", lob, state_name, channel, size_or_duration]).strip("_")


def generate_ad_name_by_account(placement_name, dimension, account, taxonomy, matched_creative=""):
    if taxonomy == "Placement Name = Ad Name":
        return str(placement_name)

    if taxonomy == "Creative Name = Ad Name":
        return matched_creative

    if taxonomy == "Dimension = Ad Name":
        return str(dimension)

    if account == "Pulte":
        return generate_pulte_ad_name(placement_name)

    if account == "Touchstone Energy":
        return generate_touchstone_ad_name(placement_name)

    if account == "Anthem / Elevance":
        return generate_elv_ad_name(placement_name, dimension)

    return str(placement_name)


def match_creatives(dimension, placement_name, creative_names):
    dim = normalize_dimension(dimension)
    placement_parts = split_placement(placement_name)
    results = []

    for creative in creative_names:
        clean_creative = normalize_text(creative).replace(" ", "").replace("*", "x")
        score = 0

        if dim and dim in clean_creative:
            score += 10

        for part in placement_parts:
            clean_part = normalize_text(part)
            if len(clean_part) >= 3 and clean_part in clean_creative:
                score += 1

        if score >= 10:
            results.append((creative, score))

    results.sort(key=lambda x: x[1], reverse=True)
    return [x[0] for x in results]


def find_utm_for_row(utm_df, placement_name, dimension):
    if utm_df.empty:
        return ""

    df = utm_df.copy()
    df.columns = [str(c).lower().strip() for c in df.columns]

    placement_col = None
    dimension_col = None
    utm_col = None

    for col in df.columns:
        if "placement" in col:
            placement_col = col
        if "dimension" in col or "size" in col:
            dimension_col = col
        if "utm" in col or "url" in col or "click" in col:
            utm_col = col

    if not utm_col:
        return ""

    placement_clean = normalize_text(placement_name)
    dim_clean = normalize_dimension(dimension)

    for _, r in df.iterrows():
        row_placement = normalize_text(r.get(placement_col, "")) if placement_col else ""
        row_dimension = normalize_dimension(r.get(dimension_col, "")) if dimension_col else ""

        placement_match = placement_clean and row_placement and (
            placement_clean in row_placement or row_placement in placement_clean
        )

        dimension_match = True
        if dimension_col:
            dimension_match = dim_clean == row_dimension

        if placement_match and dimension_match:
            return r.get(utm_col, "")

    return ""


def clear_prisma_sheet_only(ws):
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None


def paste_dataframe_to_prisma(ws, df):
    clear_prisma_sheet_only(ws)

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        if not isinstance(cell, MergedCell):
            cell.value = col_name

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if not isinstance(cell, MergedCell):
                cell.value = "" if pd.isna(value) else value


def clear_multi_sheet(ws):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None


if st.button("Generate Traffic Sheet"):

    if not os.path.exists(MASTER_TEMPLATE_PATH):
        st.error("master_template.xlsm not found. Upload it to GitHub beside Tsheet.py.")
        st.stop()

    if prisma_file is None:
        st.error("Please upload Prisma / Tsheet export.")
        st.stop()

    prisma_df = read_prisma_file(prisma_file)
    creative_names = extract_creative_names(creative_uploads)
    utm_df = read_utm_text(utm_text)

    with open(MASTER_TEMPLATE_PATH, "rb") as f:
        template_bytes = io.BytesIO(f.read())

    wb = load_workbook(template_bytes, keep_vba=True)

    ws_prisma = wb["Prisma Export - Paste as values"]
    ws_traffic = wb["Traffic_Doc"]
    ws_multi = wb["Multi-Ad or Creative Rotation"]

    paste_dataframe_to_prisma(ws_prisma, prisma_df)
    clear_multi_sheet(ws_multi)

    multi_row = 2
    review_rows = []

    for i, row in prisma_df.iterrows():
        excel_row = TRAFFIC_START_ROW + i

        try:
            dimension = row.iloc[19]        # Prisma T
            placement_name = row.iloc[20]   # Prisma U
            start_date = row.iloc[22]       # Prisma W
            end_date = row.iloc[23]         # Prisma X
        except IndexError:
            st.error("Tsheet needs columns up to X.")
            st.stop()

        dim_clean = normalize_dimension(dimension)
        matches = match_creatives(dimension, placement_name, creative_names) if creative_names else []

        matched_creative = ""
        if dim_clean == "1x1":
            matched_creative = custom_1x1
        elif matches:
            matched_creative = matches[0]

        ad_name = generate_ad_name_by_account(
            placement_name,
            dimension,
            selected_account,
            selected_taxonomy,
            matched_creative
        )

        ws_traffic[f"H{excel_row}"] = ad_name

        if matched_creative:
            ws_traffic[f"K{excel_row}"] = matched_creative
        elif creative_names:
            ws_traffic[f"K{excel_row}"] = "Creative not found"
            review_rows.append({
                "Traffic_Doc Row": excel_row,
                "Placement Name": placement_name,
                "Dimension": dimension,
                "Issue": "Creative not found"
            })

        matched_utm = find_utm_for_row(utm_df, placement_name, dimension, ad_name)
        if matched_utm:
            ws_traffic[f"{utm_output_column}{excel_row}"] = matched_utm

        if creative_type == "Multiple creatives per ad" and len(matches) > 1:
            ws_traffic[f"K{excel_row}"] = "Multiple creatives - see rotation tab"

            for creative in matches:
                ws_multi[f"A{multi_row}"] = ad_name
                ws_multi[f"D{multi_row}"] = creative
                ws_multi[f"F{multi_row}"] = "Even"
                ws_multi[f"G{multi_row}"] = start_date
                ws_multi[f"H{multi_row}"] = end_date
                multi_row += 1

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    st.success("Traffic Sheet generated successfully.")

    if creative_names:
        st.subheader("Creative Files Found")
        st.write(creative_names)

    if review_rows:
        st.warning("Some rows need manual review.")
        st.dataframe(pd.DataFrame(review_rows), use_container_width=True)

    st.download_button(
        label="Download Final Traffic Sheet",
        data=final_output,
        file_name="Final_Traffic_Sheet.xlsm",
        mime="application/vnd.ms-excel.sheet.macroEnabled.12"
    )
